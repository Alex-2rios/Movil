import io
import os
import re
from datetime import datetime, date, timedelta
from pathlib import Path

from flask import Flask, redirect, render_template, request, send_from_directory, url_for, session, flash, send_file
import psycopg
from psycopg.rows import dict_row

# reportlab
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors

# openpyxl
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

BASE_DIR = Path(__file__).resolve().parent
# Dynamically locate assets directory in the workspace
ASSETS_DIR = BASE_DIR.parent / "Copia de Terracota2p" / "assets"
if not ASSETS_DIR.exists():
    ASSETS_DIR = BASE_DIR.parent / "Terracota2p" / "assets"

app = Flask(__name__)
app.config["SECRET_KEY"] = "terracota-web-maqueta"

DATABASE_URL = os.environ.get("DATABASE_URL", "postgresql://terracota_app:TerracotaLocal123!@localhost:5433/terracota")

ROLE_MAP = {
    "Administrador": "administrador",
    "Mesero": "mesero",
    "Cocina": "cocina",
    "Cajero": "caja"
}
REV_ROLE_MAP = {
    "administrador": "Administrador",
    "mesero": "Mesero",
    "cocina": "Cocina",
    "caja": "Cajero"
}

CATEGORIA_MAP = {
    "Bebidas": "BEBIDAS",
    "Comida": "ALIMENTOS",
    "Panadería": "POSTRES",
    "Postres": "POSTRES",
    "Promociones": "PROMOS",
    "Todos": None
}

def get_db_connection():
    conn = psycopg.connect(DATABASE_URL, row_factory=dict_row)
    with conn.cursor() as cur:
        cur.execute("SET search_path TO terracota, public;")
    return conn

def slugify(text):
    text = text.lower().strip()
    text = re.sub(r'[^\w\s-]', '', text)
    text = re.sub(r'[\s_-]+', '-', text)
    return text

def parse_date(date_str, default):
    if not date_str:
        return default
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(date_str, fmt).date()
        except ValueError:
            continue
    return default

def generate_pdf_report(title, sections):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    story = []
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'TitleStyle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor("#6e321f"),
        spaceAfter=15
    )
    section_style = ParagraphStyle(
        'SectionStyle',
        parent=styles['Heading2'],
        fontSize=11,
        textColor=colors.HexColor("#73351f"),
        spaceBefore=12,
        spaceAfter=6
    )
    
    story.append(Paragraph(title, title_style))
    story.append(Spacer(1, 10))
    
    for sec in sections:
        if sec.get("title"):
            story.append(Paragraph(sec["title"], section_style))
            story.append(Spacer(1, 4))
            
        table_data = [sec["headers"]]
        for row in sec["rows"]:
            table_data.append([str(val) for val in row])
            
        t = Table(table_data)
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#cdb7a5")),
            ('TEXTCOLOR', (0,0), (-1,0), colors.HexColor("#2e211c")),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,0), 9),
            ('BOTTOMPADDING', (0,0), (-1,0), 6),
            ('BACKGROUND', (0,1), (-1,-1), colors.HexColor("#FFFDF8")),
            ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor("#eee5dc")),
            ('FONTSIZE', (0,1), (-1,-1), 8),
            ('BOTTOMPADDING', (0,1), (-1,-1), 5),
        ]))
        story.append(t)
        story.append(Spacer(1, 15))
        
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_xlsx_report(title, sections):
    buffer = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"
    
    ws.merge_cells("A1:G1")
    ws["A1"] = title
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="6E321F")
    ws["A1"].alignment = Alignment(horizontal="center")
    
    current_row = 3
    for sec in sections:
        if sec.get("title"):
            ws.cell(row=current_row, column=1, value=sec["title"]).font = Font(name="Calibri", size=12, bold=True, color="73351F")
            current_row += 1
            
        # Headers
        for col_num, header in enumerate(sec["headers"], 1):
            cell = ws.cell(row=current_row, column=col_num)
            cell.value = header
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="CDB7A5", end_color="CDB7A5", fill_type="solid")
            cell.alignment = Alignment(horizontal="left")
        current_row += 1
        
        # Rows
        for row_data in sec["rows"]:
            for col_num, val in enumerate(row_data, 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.value = val
                cell.font = Font(name="Calibri", size=10)
            current_row += 1
            
        current_row += 2  # spacing
        
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 10)
        
    wb.save(buffer)
    buffer.seek(0)
    return buffer

@app.before_request
def require_login():
    if request.endpoint in ("login", "asset", "static") or not request.endpoint:
        return
    if not session.get("logged_in"):
        return redirect(url_for("login"))

@app.context_processor
def inject_layout_data():
    return {
        "admin_user": session.get("nombre", "Admin Usuario"),
        "admin_email": session.get("usuario", "admin"),
    }

@app.get("/assets/<path:filename>")
def asset(filename):
    return send_from_directory(ASSETS_DIR, filename)

@app.route("/", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        username = request.form.get("correo", "").strip()
        password = request.form.get("password", "")
        
        if not username or not password:
            flash("Ingresa tu usuario y contraseña.", "error")
            return render_template("login.html", page_title="Iniciar Sesión", hide_nav=True)
            
        try:
            with get_db_connection() as conn:
                with conn.cursor() as cur:
                    cur.execute("SELECT * FROM terracota.autenticar_usuario(%s, %s)", (username, password))
                    user = cur.fetchone()
                    if user:
                        if "administrador" in user["roles"]:
                            session.clear()
                            session["logged_in"] = True
                            session["user_id"] = user["usuario_id"]
                            session["nombre"] = user["nombre"]
                            session["usuario"] = user["usuario"]
                            session["roles"] = list(user["roles"])
                            return redirect(url_for("inicio"))
                        else:
                            flash("Acceso denegado: Se requiere rol de Administrador.", "error")
                    else:
                        flash("Usuario o contraseña incorrectos.", "error")
        except Exception as e:
            flash(f"Error de conexión a la base de datos: {str(e)}", "error")
            
    session.clear()
    return render_template("login.html", page_title="Iniciar Sesión", hide_nav=True)

@app.get("/inicio")
def inicio():
    try:
        with get_db_connection() as conn:
            # Ventas del mes
            cur = conn.execute("""
                SELECT COALESCE(sum(total), 0) AS total 
                FROM terracota.vista_ventas_diarias 
                WHERE date_trunc('month', fecha) = date_trunc('month', now() AT TIME ZONE 'America/Mexico_City');
            """)
            ventas_mes = cur.fetchone()["total"]
            
            # Pedidos hoy
            cur = conn.execute("""
                SELECT count(*)::integer AS total 
                FROM terracota.pedidos 
                WHERE (creado_en AT TIME ZONE 'America/Mexico_City')::date = (now() AT TIME ZONE 'America/Mexico_City')::date;
            """)
            pedidos_hoy = cur.fetchone()["total"]
            
            # Usuarios activos
            cur = conn.execute("SELECT count(*)::integer AS total FROM terracota.usuarios WHERE activo;")
            usuarios_activos = cur.fetchone()["total"]
            
            # Pedidos recientes
            cur = conn.execute("""
                SELECT p.id, m.numero AS mesa, u.nombre AS mesero, p.estado, p.total, to_char(p.creado_en, 'HH24:MI') AS hora
                FROM terracota.pedidos p
                JOIN terracota.mesas m ON m.id = p.mesa_id
                JOIN terracota.usuarios u ON u.id = p.mesero_id
                ORDER BY p.creado_en DESC
                LIMIT 10;
            """)
            pedidos = cur.fetchall()
            
            # Datos gráfica: Ingresos vs Gastos últimos 7 días
            cur = conn.execute("""
                WITH fechas AS (
                  SELECT generate_series(
                    (now() AT TIME ZONE 'America/Mexico_City')::date - interval '6 days',
                    (now() AT TIME ZONE 'America/Mexico_City')::date,
                    '1 day'::interval
                  )::date AS fecha
                )
                SELECT
                  f.fecha,
                  CASE EXTRACT(DOW FROM f.fecha)
                    WHEN 0 THEN 'Dom' WHEN 1 THEN 'Lun' WHEN 2 THEN 'Mar'
                    WHEN 3 THEN 'Mié' WHEN 4 THEN 'Jue' WHEN 5 THEN 'Vie'
                    WHEN 6 THEN 'Sáb' END AS dia_semana,
                  COALESCE((SELECT sum(total) FROM terracota.vista_ventas_diarias v WHERE v.fecha = f.fecha), 0)::numeric(12,2) AS ganancias,
                  COALESCE((SELECT sum(monto) FROM terracota.gastos g WHERE g.fecha = f.fecha), 0)::numeric(12,2) AS gastos
                FROM fechas f
                ORDER BY f.fecha;
            """)
            chart_sales = cur.fetchall()
            
            # Datos gráfica: Productos más vendidos
            cur = conn.execute("""
                SELECT nombre_producto, sum(cantidad)::integer AS total_vendido
                FROM terracota.pedido_detalles pd
                JOIN terracota.pedidos p ON p.id = pd.pedido_id
                GROUP BY nombre_producto
                ORDER BY total_vendido DESC
                LIMIT 5;
            """)
            chart_products = cur.fetchall()
            
    except Exception as e:
        flash(f"Error cargando estadísticas: {str(e)}", "error")
        ventas_mes = 0
        pedidos_hoy = 0
        usuarios_activos = 0
        pedidos = []
        chart_sales = []
        chart_products = []
        
    metrics = [
        {"icon": "Ventas del Mes.png", "label": "Ventas del Mes", "value": f"${ventas_mes:,.2f}"},
        {"icon": "Pedidos de Hoy.png", "label": "Pedidos Hoy", "value": str(pedidos_hoy)},
        {"icon": "Usuario.png", "label": "Usuarios Activos", "value": str(usuarios_activos)},
        {"icon": "Crecimiento.png", "label": "Estado", "value": "Estable"},
    ]
    
    return render_template(
        "inicio.html",
        page_title="Estadísticas",
        active_page="inicio",
        pedidos=pedidos,
        metrics=metrics,
        chart_sales=chart_sales,
        chart_products=chart_products
    )

@app.get("/estadisticas")
def estadisticas():
    return inicio()

@app.get("/usuarios")
def usuarios():
    try:
        with get_db_connection() as conn:
            cur = conn.execute("""
                SELECT u.id, u.nombre, u.usuario, u.activo,
                       COALESCE(array_agg(r.clave ORDER BY r.clave)
                         FILTER (WHERE r.clave IS NOT NULL), ARRAY[]::varchar[]) AS roles,
                       CASE WHEN u.activo THEN 'Activo' ELSE 'Inactivo' END AS estado,
                       to_char(u.actualizado_en, 'YYYY-MM-DD HH24:MI') AS ultimo_acceso
                FROM terracota.usuarios u
                LEFT JOIN terracota.usuario_roles ur ON ur.usuario_id = u.id
                LEFT JOIN terracota.roles r ON r.id = ur.rol_id
                GROUP BY u.id
                ORDER BY u.nombre;
            """)
            db_users = cur.fetchall()
            
            usuarios_mapped = []
            for u in db_users:
                rol_list = [REV_ROLE_MAP.get(r, r.capitalize()) for r in u["roles"]]
                rol_str = rol_list[0] if rol_list else "Mesero"
                
                usuarios_mapped.append({
                    "id": u["id"],
                    "nombre": u["nombre"],
                    "email": u["usuario"],
                    "telefono": "-",
                    "rol": rol_str,
                    "estado": u["estado"],
                    "ultimo_acceso": u["ultimo_acceso"] or "Nunca"
                })
    except Exception as e:
        flash(f"Error cargando usuarios: {str(e)}", "error")
        usuarios_mapped = []
        
    return render_template(
        "usuarios.html",
        page_title="Gestión de Usuarios",
        active_page="usuarios",
        usuarios=usuarios_mapped,
    )

@app.route("/usuarios/nuevo", methods=["GET", "POST"])
def agregar_usuario():
    if request.method == "POST":
        nombre = request.form.get("nombre", "").strip()
        email = request.form.get("email", "").strip()
        rol_display = request.form.get("rol", "Mesero")
        password = request.form.get("password", "")
        confirmacion = request.form.get("confirmacion", "")
        
        if not nombre or not email or not password:
            flash("Todos los campos son obligatorios.", "error")
            return redirect(url_for("agregar_usuario"))
            
        if password != confirmacion:
            flash("La contraseña y la confirmación no coinciden.", "error")
            return redirect(url_for("agregar_usuario"))
            
        rol_key = ROLE_MAP.get(rol_display, "mesero")
        
        try:
            with get_db_connection() as conn:
                cur = conn.execute("SELECT id FROM terracota.usuarios WHERE lower(usuario) = lower(%s)", (email,))
                if cur.fetchone():
                    flash(f"El usuario '{email}' ya existe.", "error")
                    return redirect(url_for("agregar_usuario"))
                    
                cur = conn.execute("""
                    INSERT INTO terracota.usuarios(nombre, usuario, password_hash)
                    VALUES (%s, %s, crypt(%s, gen_salt('bf')))
                    RETURNING id;
                """, (nombre, email, password))
                user_id = cur.fetchone()["id"]
                
                cur_rol = conn.execute("SELECT id FROM terracota.roles WHERE clave = %s", (rol_key,))
                rol_record = cur_rol.fetchone()
                if rol_record:
                    conn.execute("INSERT INTO terracota.usuario_roles(usuario_id, rol_id) VALUES (%s, %s)", (user_id, rol_record["id"]))
                
                conn.commit()
                flash("Usuario creado con éxito.", "success")
                return redirect(url_for("usuarios"))
        except Exception as e:
            flash(f"Error creando usuario: {str(e)}", "error")
            return redirect(url_for("agregar_usuario"))
            
    usuario_template = {
        "id": 0,
        "nombre": "",
        "email": "",
        "telefono": "",
        "rol": "Mesero",
    }
    return render_template(
        "usuario_form.html",
        page_title="Agregar Usuario",
        active_page="usuarios",
        action_label="Agregar Usuario",
        usuario=usuario_template,
    )

@app.route("/usuarios/<int:user_id>/editar", methods=["GET", "POST"])
def editar_usuario(user_id):
    if request.method == "POST":
        nombre = request.form.get("nombre", "").strip()
        email = request.form.get("email", "").strip()
        rol_display = request.form.get("rol", "Mesero")
        password = request.form.get("password", "")
        confirmacion = request.form.get("confirmacion", "")
        
        if not nombre or not email:
            flash("Nombre y Email son obligatorios.", "error")
            return redirect(url_for("editar_usuario", user_id=user_id))
            
        rol_key = ROLE_MAP.get(rol_display, "mesero")
        
        try:
            with get_db_connection() as conn:
                conn.execute("""
                    UPDATE terracota.usuarios
                    SET nombre = %s, usuario = %s, actualizado_en = now()
                    WHERE id = %s;
                """, (nombre, email, user_id))
                
                if password and password != "12345678":
                    if password != confirmacion:
                        flash("La contraseña y la confirmación no coinciden.", "error")
                        return redirect(url_for("editar_usuario", user_id=user_id))
                    conn.execute("""
                        UPDATE terracota.usuarios
                        SET password_hash = crypt(%s, gen_salt('bf'))
                        WHERE id = %s;
                    """, (password, user_id))
                    
                cur_rol = conn.execute("SELECT id FROM terracota.roles WHERE clave = %s", (rol_key,))
                rol_record = cur_rol.fetchone()
                if rol_record:
                    conn.execute("DELETE FROM terracota.usuario_roles WHERE usuario_id = %s", (user_id,))
                    conn.execute("INSERT INTO terracota.usuario_roles(usuario_id, rol_id) VALUES (%s, %s)", (user_id, rol_record["id"]))
                
                conn.commit()
                flash("Usuario actualizado con éxito.", "success")
                return redirect(url_for("usuarios"))
        except Exception as e:
            flash(f"Error actualizando usuario: {str(e)}", "error")
            return redirect(url_for("editar_usuario", user_id=user_id))
            
    try:
        with get_db_connection() as conn:
            cur = conn.execute("""
                SELECT u.id, u.nombre, u.usuario,
                       COALESCE(array_agg(r.clave ORDER BY r.clave)
                         FILTER (WHERE r.clave IS NOT NULL), ARRAY[]::varchar[]) AS roles
                FROM terracota.usuarios u
                LEFT JOIN terracota.usuario_roles ur ON ur.usuario_id = u.id
                LEFT JOIN terracota.roles r ON r.id = ur.rol_id
                WHERE u.id = %s
                GROUP BY u.id;
            """, (user_id,))
            u = cur.fetchone()
            if u:
                rol_list = [REV_ROLE_MAP.get(r, r.capitalize()) for r in u["roles"]]
                rol_str = rol_list[0] if rol_list else "Mesero"
                usuario_mapped = {
                    "id": u["id"],
                    "nombre": u["nombre"],
                    "email": u["usuario"],
                    "telefono": "-",
                    "rol": rol_str,
                }
            else:
                flash("Usuario no encontrado.", "error")
                return redirect(url_for("usuarios"))
    except Exception as e:
        flash(f"Error cargando usuario: {str(e)}", "error")
        return redirect(url_for("usuarios"))
        
    return render_template(
        "usuario_form.html",
        page_title="Editar Usuario",
        active_page="usuarios",
        action_label="Actualizar Usuario",
        usuario=usuario_mapped,
    )

@app.post("/usuarios/<int:user_id>/eliminar")
def eliminar_usuario(user_id):
    if user_id == session.get("user_id"):
        flash("No puedes desactivar tu propia cuenta.", "error")
        return redirect(url_for("usuarios"))
        
    try:
        with get_db_connection() as conn:
            conn.execute("UPDATE terracota.usuarios SET activo = false WHERE id = %s;", (user_id,))
            conn.commit()
            flash("Usuario desactivado con éxito.", "success")
    except Exception as e:
        flash(f"Error desactivando usuario: {str(e)}", "error")
        
    return redirect(url_for("usuarios"))

@app.get("/reportes")
def reportes():
    return render_template("reportes.html", page_title="Reportes", active_page="reportes")

@app.post("/reportes/exportar")
def exportar_reporte():
    tipo = request.form.get("tipo_reporte", "ventas")
    fecha_inicio_str = request.form.get("fecha_inicio", "")
    fecha_fin_str = request.form.get("fecha_fin", "")
    categoria = request.form.get("categoria", "Todos")
    estado = request.form.get("estado", "Todos")
    formato = request.form.get("formato", "pdf").lower()
    
    fecha_inicio = parse_date(fecha_inicio_str, date(2026, 1, 1))
    fecha_fin = parse_date(fecha_fin_str, date.today() + timedelta(days=1))
    
    sections = []
    try:
        with get_db_connection() as conn:
            if tipo == "ventas":
                # 1. Desglose de Ventas por Producto
                cat_cond = ""
                params_prod = [fecha_inicio, fecha_fin]
                if categoria != "Todos los productos" and categoria != "Todos":
                    cat_cond = " AND c.nombre = %s"
                    params_prod.append(categoria.upper())
                    
                query_prod = f"""
                    SELECT pd.nombre_producto, c.nombre AS categoria, max(pr.stock_actual)::integer AS stock_actual, sum(pd.cantidad)::integer AS cantidad_vendida, sum(pd.importe)::numeric(12,2) AS total_recaudado
                    FROM terracota.pedido_detalles pd
                    JOIN terracota.pedidos p ON p.id = pd.pedido_id
                    JOIN terracota.productos pr ON pr.id = pd.producto_id
                    JOIN terracota.categorias c ON c.id = pr.categoria_id
                    WHERE p.creado_en::date BETWEEN %s AND %s{cat_cond}
                    GROUP BY pd.nombre_producto, c.nombre
                    ORDER BY cantidad_vendida DESC;
                """
                cur_prod = conn.execute(query_prod, params_prod)
                records_prod = cur_prod.fetchall()
                rows_prod = [[r["nombre_producto"], r["categoria"], r["stock_actual"], r["cantidad_vendida"], f"${float(r['total_recaudado']):,.2f}"] for r in records_prod]
                
                sections.append({
                    "title": "1. Resumen de Ventas por Producto (Volumen y Recaudación)",
                    "headers": ["Producto", "Categoria", "Stock Actual", "Cantidad Vendida", "Total Recaudado"],
                    "rows": rows_prod
                })

                # 2. Historial de Transacciones
                estado_cond = ""
                params = [fecha_inicio, fecha_fin]
                if estado != "Todos":
                    estado_cond = " AND p.estado = %s"
                    params.append(estado.upper())
                    
                query = f"""
                    SELECT pg.id AS pago_id, t.folio, m.numero AS mesa, u.nombre AS mesero, pg.metodo, pg.total, 
                           to_char(pg.creado_en, 'YYYY-MM-DD HH24:MI') AS fecha
                    FROM terracota.pagos pg
                    JOIN terracota.pedidos p ON p.id = pg.pedido_id
                    JOIN terracota.mesas m ON m.id = p.mesa_id
                    JOIN terracota.usuarios u ON u.id = p.mesero_id
                    LEFT JOIN terracota.tickets t ON t.pago_id = pg.id
                    WHERE pg.creado_en::date BETWEEN %s AND %s{estado_cond}
                    ORDER BY pg.creado_en DESC;
                """
                cur = conn.execute(query, params)
                records = cur.fetchall()
                rows = [[r["pago_id"], r["folio"] or "-", r["mesa"], r["mesero"], r["metodo"], f"${float(r['total']):,.2f}", r["fecha"]] for r in records]
                
                sections.append({
                    "title": "2. Detalle de Transacciones (Pedidos Pagados)",
                    "headers": ["ID Pago", "Folio Ticket", "Mesa", "Mesero", "Metodo", "Total", "Fecha"],
                    "rows": rows
                })
                
                title = f"Reporte Consolidado de Ventas ({fecha_inicio} a {fecha_fin})"
                
            elif tipo == "usuarios":
                estado_cond = ""
                params = []
                if estado != "Todos":
                    estado_cond = " WHERE u.activo = %s"
                    params.append(estado == "Activo")
                    
                query = f"""
                    SELECT u.nombre, u.usuario, CASE WHEN u.activo THEN 'Activo' ELSE 'Inactivo' END AS estado,
                           COALESCE(array_to_string(array_agg(r.nombre), ', '), 'Ninguno') AS roles,
                           to_char(u.creado_en, 'YYYY-MM-DD') AS fecha_registro
                    FROM terracota.usuarios u
                    LEFT JOIN terracota.usuario_roles ur ON ur.usuario_id = u.id
                    LEFT JOIN terracota.roles r ON r.id = ur.rol_id
                    {estado_cond}
                    GROUP BY u.id
                    ORDER BY u.nombre;
                """
                cur = conn.execute(query, params)
                records = cur.fetchall()
                rows = [[r["nombre"], r["usuario"], r["estado"], r["roles"], r["fecha_registro"]] for r in records]
                
                sections.append({
                    "headers": ["Nombre", "Usuario", "Estado", "Roles", "Fecha Registro"],
                    "rows": rows
                })
                
                title = "Reporte de Usuarios del Sistema"
                
            elif tipo == "inventario":
                params = []
                cat_cond = ""
                if categoria != "Inventario general" and categoria != "Todos":
                    cat_cond = " WHERE c.nombre = %s"
                    params.append(categoria.upper())
                    
                query = f"""
                    SELECT p.nombre, c.nombre AS categoria, p.stock_actual, p.stock_minimo, p.precio,
                           CASE WHEN p.stock_actual <= p.stock_minimo THEN 'Bajo stock' ELSE 'Disponible' END AS estado
                    FROM terracota.productos p
                    JOIN terracota.categorias c ON c.id = p.categoria_id
                    {cat_cond}
                    ORDER BY c.nombre, p.nombre;
                """
                cur = conn.execute(query, params)
                records = cur.fetchall()
                rows = [[r["nombre"], r["categoria"], r["stock_actual"], r["stock_minimo"], f"${float(r['precio']):,.2f}", r["estado"]] for r in records]
                
                sections.append({
                    "headers": ["Producto", "Categoria", "Stock Actual", "Stock Minimo", "Precio Unitario", "Estado"],
                    "rows": rows
                })
                
                title = "Reporte de Inventario y Existencias"
                
            else:
                flash("Tipo de reporte no reconocido.", "error")
                return redirect(url_for("reportes"))
                
    except Exception as e:
        flash(f"Error generando reporte: {str(e)}", "error")
        return redirect(url_for("reportes"))
        
    if formato == "pdf":
        buffer = generate_pdf_report(title, sections)
        filename = f"reporte_{tipo}_{datetime.now().strftime('%Y%m%d')}.pdf"
        return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/pdf")
    else:
        buffer = generate_xlsx_report(title, sections)
        filename = f"reporte_{tipo}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        return send_file(buffer, as_attachment=True, download_name=filename, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@app.get("/inventario")
def inventario():
    try:
        with get_db_connection() as conn:
            cur = conn.execute("""
                SELECT p.id, p.nombre AS producto, c.nombre AS categoria, p.stock_actual, p.stock_minimo, p.precio,
                       CASE WHEN p.stock_actual <= p.stock_minimo THEN 'Bajo' ELSE 'Disponible' END AS estado
                FROM terracota.productos p
                JOIN terracota.categorias c ON c.id = p.categoria_id
                ORDER BY p.nombre;
            """)
            productos = cur.fetchall()
            
            productos_mapped = []
            alertas = []
            for p in productos:
                prod = {
                    "id": p["id"],
                    "producto": p["producto"],
                    "categoria": p["categoria"].lower(),
                    "stock_actual": p["stock_actual"],
                    "stock_minimo": p["stock_minimo"],
                    "precio": f"${float(p["precio"]):,.2f}",
                    "estado": p["estado"]
                }
                productos_mapped.append(prod)
                if p["estado"] == "Bajo":
                    alertas.append(prod)
    except Exception as e:
        flash(f"Error cargando inventario: {str(e)}", "error")
        productos_mapped = []
        alertas = []
        
    return render_template(
        "inventario.html",
        page_title="Gestión de Inventario",
        active_page="inventario",
        productos=productos_mapped,
        alertas=alertas[:2],
    )

@app.route("/inventario/nuevo", methods=["GET", "POST"])
def agregar_producto():
    if request.method == "POST":
        nombre = request.form.get("producto", "").strip()
        categoria_display = request.form.get("categoria", "Bebidas")
        stock_actual = int(request.form.get("stock_actual", "50"))
        stock_minimo = int(request.form.get("stock_minimo", "15"))
        precio_str = request.form.get("precio", "0").replace("$", "").strip()
        precio = float(precio_str)
        
        if not nombre:
            flash("El nombre del producto es obligatorio.", "error")
            return redirect(url_for("agregar_producto"))
            
        clave = slugify(nombre)
        categoria_clave = CATEGORIA_MAP.get(categoria_display, "BEBIDAS")
        
        try:
            with get_db_connection() as conn:
                cur_cat = conn.execute("SELECT id FROM terracota.categorias WHERE clave = %s", (categoria_clave,))
                cat_record = cur_cat.fetchone()
                if not cat_record:
                    flash("Categoría inválida.", "error")
                    return redirect(url_for("agregar_producto"))
                    
                cur = conn.execute("SELECT id FROM terracota.productos WHERE clave = %s", (clave,))
                if cur.fetchone():
                    flash(f"El producto con clave '{clave}' ya existe.", "error")
                    return redirect(url_for("agregar_producto"))
                    
                conn.execute("""
                    INSERT INTO terracota.productos(clave, categoria_id, nombre, precio, stock_actual, stock_minimo)
                    VALUES (%s, %s, %s, %s, %s, %s);
                """, (clave, cat_record["id"], nombre, precio, stock_actual, stock_minimo))
                conn.commit()
                flash("Producto creado con éxito.", "success")
                return redirect(url_for("inventario"))
        except Exception as e:
            flash(f"Error creando producto: {str(e)}", "error")
            return redirect(url_for("agregar_producto"))
            
    producto_template = {
        "id": 0,
        "producto": "",
        "categoria": "bebidas",
        "stock_actual": "",
        "stock_minimo": "",
        "precio": "",
    }
    return render_template(
        "producto_form.html",
        page_title="Agregar Producto",
        active_page="inventario",
        action_label="Agregar Producto",
        producto=producto_template,
    )

@app.route("/inventario/<int:product_id>/editar", methods=["GET", "POST"])
def editar_producto(product_id):
    if request.method == "POST":
        nombre = request.form.get("producto", "").strip()
        categoria_display = request.form.get("categoria", "Bebidas")
        stock_actual = int(request.form.get("stock_actual", "50"))
        stock_minimo = int(request.form.get("stock_minimo", "15"))
        precio_str = request.form.get("precio", "0").replace("$", "").strip()
        precio = float(precio_str)
        
        if not nombre:
            flash("El nombre del producto es obligatorio.", "error")
            return redirect(url_for("editar_producto", product_id=product_id))
            
        categoria_clave = CATEGORIA_MAP.get(categoria_display, "BEBIDAS")
        
        try:
            with get_db_connection() as conn:
                cur_cat = conn.execute("SELECT id FROM terracota.categorias WHERE clave = %s", (categoria_clave,))
                cat_record = cur_cat.fetchone()
                if not cat_record:
                    flash("Categoría inválida.", "error")
                    return redirect(url_for("editar_producto", product_id=product_id))
                    
                conn.execute("""
                    UPDATE terracota.productos
                    SET nombre = %s, categoria_id = %s, stock_actual = %s, stock_minimo = %s, precio = %s, actualizado_en = now()
                    WHERE id = %s;
                """, (nombre, cat_record["id"], stock_actual, stock_minimo, precio, product_id))
                conn.commit()
                flash("Producto actualizado con éxito.", "success")
                return redirect(url_for("inventario"))
        except Exception as e:
            flash(f"Error actualizando producto: {str(e)}", "error")
            return redirect(url_for("editar_producto", product_id=product_id))
            
    try:
        with get_db_connection() as conn:
            cur = conn.execute("""
                SELECT p.id, p.nombre, c.clave AS categoria_clave, p.stock_actual, p.stock_minimo, p.precio
                FROM terracota.productos p
                JOIN terracota.categorias c ON c.id = p.categoria_id
                WHERE p.id = %s;
            """, (product_id,))
            p = cur.fetchone()
            if p:
                categoria_display = "bebidas"
                for k, v in CATEGORIA_MAP.items():
                    if v == p["categoria_clave"]:
                        categoria_display = k.lower()
                        break
                        
                producto_mapped = {
                    "id": p["id"],
                    "producto": p["nombre"],
                    "categoria": categoria_display,
                    "stock_actual": p["stock_actual"],
                    "stock_minimo": p["stock_minimo"],
                    "precio": float(p["precio"]),
                }
            else:
                flash("Producto no encontrado.", "error")
                return redirect(url_for("inventario"))
    except Exception as e:
        flash(f"Error cargando producto: {str(e)}", "error")
        return redirect(url_for("inventario"))
        
    return render_template(
        "producto_form.html",
        page_title="Editar Producto",
        active_page="inventario",
        action_label="Editar Producto",
        producto=producto_mapped,
    )

@app.post("/accion/simulada")
def accion_simulada():
    destino = request.form.get("destino") or "inicio"
    return redirect(url_for(destino))

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
